import openpyxl
car_file =openpyxl.load_workbook("Car_Rantal.xlsx")

car_list =car_file['Tabelle1']
#how many type car is available
car_per_name = {}
car_per_passagiere={}
car_price_type = {}
dicount_car={}



for car_row in range(2, car_list.max_row + 1):
    number_passengers = car_list.cell(car_row, 3).value
    preis_car=car_list.cell(car_row,7).value
    type_car =car_list.cell(car_row,5).value




   # print(type_car)
    if type_car in car_per_name:
        current_name_product= car_per_name[type_car]
        car_per_name[type_car]=current_name_product +1

# calculate type with number of passengers
        car_per_passagiere[type_car]=number_passengers
    else:
        car_per_name[type_car] =1
    if type_car in car_price_type:
        current_total_price =car_price_type.get(type_car)
        car_price_type[type_car]=(preis_car)
        dicount_car[type_car]=0.3*(preis_car)
    car_price_type[type_car]=(preis_car)
    dicount_car[type_car]=round(0.3 *(preis_car))
class User:


    def __init__(self,name, user_email, password,numberOffamily,select_car_user):
        self.name = name
        self.email=user_email
        self.password=password
        self.numberOffamily=numberOffamily
        self.id_booking = 0
        self.select_car_user = select_car_user
        self.offer_auto = ""
        self.price = 0
        self.discount=0
        self.anzahle_auto=0
        self.gesamt_cost=0

    def herzlich_welcome(self):
        print("Herzlichen welcomen in Car rental\n")

        print(f"see available taxis at the moment with copacity: {car_per_passagiere}\n")
        print(f"Car rental price in one day {car_price_type} Euro\n")
        print(f"if the car is reserved as a family, we have a 30% discount for you: {dicount_car} Euro \n")

    def get_number_family(self):
        for i in range(1):
            family=input("are you family: yes oder no: ").lower()

            if family=="no":


                break
            else:
                try:
                    numberOf_family=int(input("how many people did you consist of: "))
                    self.numberOffamily = numberOf_family
                except ValueError:
                    print("you can to write just string")
                    numberOf_family = int(input("how many people did you consist of: "))
                    self.numberOffamily = numberOf_family



    def change_password(self,new_password):
        self.password=new_password
    #def get_select_car(self,which_auto):
     #   self.select_car = which_auto




    def get_offer_auto(self):
        while 1:
            if self.numberOffamily >= 2:
                offer = input("I can offer you Hachback or SUV: please select your car otherweise write no: ").lower()
                if offer == "hachback" or offer == "suv":
                    self.offer_auto = offer
                    break

            else:
                offer = input("I can offer you Luxy or SUV: please select your car otherweise write no: ").lower()
                self.offer_auto = offer
                break

        return self.offer_auto


    def get_select_car(self):
         #neu_select_car_user = input(" Please choose the car you want: ")
        self.select_car_user = self.offer_auto
        return self.select_car_user


    def get_calculate_price(self):
        if self.offer_auto=="hachback":
            price1 =car_price_type['Hatchback']
        elif self.offer_auto=="suv":
            price1 = car_price_type["SUV"]
        elif self.offer_auto=="sendan":
            price1=car_price_type["Hatchback"]
        else:
            price1=car_price_type["Sendan"]
        self.price=price1

        return self.price*self.anzahle_auto

    def get_anzahle_auto(self):
        if (4<=self.numberOffamily<=8) and (self.offer_auto == "hachback" or self.offer_auto == "sendan"):
            self.anzahle_auto = 2
            return self.anzahle_auto
        elif (self.numberOffamily<= 4) and (self.offer_auto == "hachback" or self.offer_auto == "sendan"):
            self.anzahle_auto = 1
            return self.anzahle_auto
        elif (8<self.numberOffamily<=12) and (self.offer_auto == "hachback" or self.offer_auto == "sendan"):
            self.anzahle_auto = 3
            return self.anzahle_auto
        elif (2<=self.numberOffamily<=8) and (self.offer_auto == "suv"):
            self.anzahle_auto = 1
            return self.anzahle_auto
        elif (8<self.numberOffamily<=16) and (self.offer_auto == "suv"):
            self.anzahle_auto = 2
            return self.anzahle_auto
        else:
            self.anzahle_auto = 1
            return self.anzahle_auto

    def get_calculate_auto_rabbatt(self):
        if self.numberOffamily >= 2:
            self.discount = (self.price * 0.3) - (self.price * self.anzahle_auto)
        else:
            self.discount =0
        return self.discount



    def gesamt_preice(self):
        self.gesamt_cost=(self.price -self.discount) *self.anzahle_auto

    def get_user_info(self):
        print(f"User {self.name} current nummber of family is {self.numberOffamily}\n"   
        f" and chose auto is {self.offer_auto}\n the daily price of each car is {self.price} Euro and mit discount {self.discount} Euro, per days! \n"
        f" total cost is {self.gesamt_cost} "
        f"and anzahle des Auto ist {self.anzahle_auto}")








    # def arrange_booking(self,anzahl_ auto):









